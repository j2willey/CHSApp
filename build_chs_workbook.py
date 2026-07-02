#!/usr/bin/env python3
"""
CHS Scheduling Workbook Builder + Firestore Push
Accepts Event_Data_Dump*.csv and/or myEventClassData*.csv.
Multiple files merged. Auto-detects format.

Usage:
  python3 build_chs_workbook.py <file.csv> [file2.csv ...] [output_dir]
  python3 build_chs_workbook.py <file.csv> --service-account <key.json>   # also push to Firestore
  python3 build_chs_workbook.py <file.csv> --push-only <key.json>          # Firestore only, no xlsx
"""
import csv, re, sys, os, math
from collections import Counter, defaultdict
from datetime import date, datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run([sys.executable,'-m','pip','install','openpyxl','--break-system-packages','-q'])
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

# ── COLORS ──────────────────────────────────────────────────────────────────
BLUE="003F87"; RED="CE1126"; TAN="D6CEBD"; GRAY="515354"
MED_BLUE="2E75B6"; DK_GREEN="1F7A4D"; ORANGE="ED7D31"; PURPLE="7030A0"
LT_GRAY_BG="F5F5F5"; LT_GREEN="E2EFDA"; LT_YELLOW="FFF2CC"
LT_RED="FCE4D6"; LT_GRAY="F0F0F0"; WHITE="FFFFFF"; ALT_GRAY="F2F2F2"

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, italic=False, size=10, color="000000", name="Arial"):
    return Font(bold=bold, italic=italic, size=size, color=color, name=name)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
def hdr(ws, rn, vals, bg, fg="FFFFFF", bold=True, size=10, wrap=False, height=None):
    for ci,val in enumerate(vals,1):
        c=ws.cell(row=rn,column=ci,value=val)
        c.fill=fill(bg); c.font=font(bold=bold,color=fg,size=size); c.alignment=align(wrap=wrap)
    if height: ws.row_dimensions[rn].height=height

# ── FORMAT DETECTION & PARSING ───────────────────────────────────────────────
YOUTH_TYPES = {
    'Scouts BSA Youth Before November 15th','Scouts BSA Youth- In Camp',
    'Scouts BSA Youth after November 15th','CIT - Counselor in Training',
    'SPL Free Week','Youth Solo Camper','SPL Paid','Senior Patrol Leader (SPL)',
}

def detect_format(headers):
    if 'Registration Number' in headers: return 'A'
    if 'Block 1 Room' in headers: return 'B'
    raise ValueError("Unrecognized CSV — expected Event_Data_Dump or myEventClassData columns.")

class Scout:
    __slots__ = ('first','last','unit','district','council','reg_num','rank','tshirt',
                 'balance','cd_email','cd_pct','parent_email','parent_phone',
                 'admin_comment','blocks','rooms','registrant_type','session')
    def __init__(self, **kw):
        for k in self.__slots__: setattr(self, k, kw.get(k,''))
        if not self.blocks: self.blocks=['']*6
        if not self.rooms:  self.rooms=['']*6
    @property
    def name(self): return f"{self.last}, {self.first}"
    @property
    def full_name(self): return f"{self.first} {self.last}"

def parse_fmt_a(rows):
    scouts=[]
    for r in rows:
        if not r['First Name'].strip(): continue
        if r['Registrant Type'] not in YOUTH_TYPES: continue
        if r.get('Cancelled','No')=='Yes': continue
        try: balance=float(r.get('Balance Due') or 0)
        except ValueError: balance=0.0
        tshirt=''
        for s in ['Adult Small','Adult Medium','Adult Large','Adult X-Large',
                  'Adult 2X-Large','Adult 3X-Large','Adult 4X-Large']:
            if r.get(f'Camp T-Shirt Free T Shirt - {s}','')=='1':
                tshirt=s.replace('Adult ',''); break
        scouts.append(Scout(
            first=r['First Name'].strip(), last=r['Last Name'].strip(),
            unit=r.get('Unit',''), district=r.get('District',''),
            council=r.get('Council',''), reg_num=r.get('Registration Number',''),
            rank='', tshirt=tshirt, balance=balance,
            cd_email=r.get('CampDoc Email','').strip(),
            cd_pct=r.get('CampDoc Completeness',''),
            parent_email=r.get('Parent Guardian Email',''),
            parent_phone=r.get('Parent Guardian Phone',''),
            admin_comment=r.get('Admin Comment',''),
            blocks=[r.get(f'Block {i}','') for i in range(1,7)],
            rooms=['']*6, registrant_type=r.get('Registrant Type',''),
            session=r.get('Session',''),
        ))
    return scouts

def parse_fmt_b(rows):
    scouts=[]
    for r in rows:
        if not r['First Name'].strip(): continue
        if r['Registrant Type'] not in YOUTH_TYPES: continue
        unit=f"{r.get('Unit Type','')} {r.get('Unit Nbr.','')} {r.get('Unit Designation','')}".strip()
        scouts.append(Scout(
            first=r['First Name'].strip(), last=r['Last Name'].strip(),
            unit=unit, district=r.get('District Name',''),
            council=r.get('Council Name',''), reg_num=r.get('Reg Nbr.',''),
            rank=r.get('Rank',''), tshirt='', balance=0.0,
            cd_email='', cd_pct='',
            parent_email=r.get('Email',''), parent_phone='', admin_comment='',
            blocks=[r.get(f'Block {i}','') for i in range(1,7)],
            rooms=[r.get(f'Block {i} Room','') for i in range(1,7)],
            registrant_type=r.get('Registrant Type',''),
            session=r.get('Session',''),
        ))
    return scouts

# ── BADGE NORMALIZATION ──────────────────────────────────────────────────────
EAGLE_REQUIRED={
    "First Aid","Citizenship in the Community","Citizenship in the Nation",
    "Citizenship in the World","Communications","Communication","Cooking",
    "Personal Fitness","Emergency Preparedness","Lifesaving",
    "Environmental Science","Sustainability","Personal Management",
    "Swimming","Cycling","Hiking & Backpacking","Camping","Family Life",
}
TWO_PERIOD={"Archery","Climbing","Rifle Shooting","Shotgun Shooting","Welding"}

def normalize_mb_cell(raw):
    if not raw or not raw.strip(): return []
    raw=raw.strip()
    if any(p in raw.lower() for p in ['(no badge)','spl 101']): return []
    raw=re.sub(r'\s*\((?:Mo|Tu|We|Th|Fr|Sa|Su)+\)\s*$','',raw)
    parts=re.split(r',\s+(?=[A-Z])',raw)
    result=[]
    for part in parts:
        b=part.strip()
        b=re.sub(r'\s*\(\d{4}\s+version\)','',b)
        b=re.sub(r'\s*-\s*Qualify\s+\d+','',b)
        b=re.sub(r'\s*[-]\s*(Morning|Afternoon)\s*$','',b,flags=re.I)
        if re.search(r'citizenship in (the )?society',b,re.I): b='Citizenship in Society (DISCONTINUED)'
        if re.search(r'backpacking',b,re.I) and re.search(r'hiking',b,re.I): b='Hiking & Backpacking'
        b=b.replace('Substainablity','Sustainability').strip()
        if b: result.append(b)
    return result

def get_scout_mbs(s):
    seen=set(); result=[]
    for raw in s.blocks:
        for mb in normalize_mb_cell(raw):
            if mb not in seen: seen.add(mb); result.append(mb)
    return result

def is_eagle(mb): return mb in EAGLE_REQUIRED or mb.replace('Communication','Communications') in EAGLE_REQUIRED
def is_two_period(mb): return any(tp.lower() in mb.lower() for tp in TWO_PERIOD)
def is_discontinued(mb): return 'DISCONTINUED' in mb

# ── DASHBOARD ────────────────────────────────────────────────────────────────
def build_dashboard(wb, scouts, session_label, fmt):
    ws=wb.create_sheet("Dashboard"); ws.sheet_properties.tabColor=BLUE
    total=len(scouts); svmbc=sum(1 for s in scouts if 'Silicon Valley' in s.council); ooc=total-svmbc
    bal_count=sum(1 for s in scouts if s.balance>0)
    complete=sum(1 for s in scouts if s.cd_pct=='100')
    zero_cd=sum(1 for s in scouts if s.cd_pct=='0' and s.cd_email)
    disc_count=len({s.reg_num for s in scouts if any(is_discontinued(mb) for mb in get_scout_mbs(s))})
    tshirt_count=sum(1 for s in scouts if s.tshirt)
    today_str=date.today().strftime("%B %d, %Y")

    ws.merge_cells('A1:L1'); c=ws['A1']
    c.value="Camp Hi-Sierra 2026  |  Scheduling Workbook"
    c.fill=fill(BLUE); c.font=font(bold=True,color=WHITE,size=16); c.alignment=align(h="center"); ws.row_dimensions[1].height=30

    ws.merge_cells('A2:L2'); c=ws['A2']
    c.value=f"{session_label}  |  Generated {today_str}  |  Source: {'Event Data Dump' if fmt=='A' else 'Class Schedule Export'}"
    c.fill=fill(GRAY); c.font=font(color=WHITE,size=11); c.alignment=align(h="center"); ws.row_dimensions[2].height=20
    ws.row_dimensions[3].height=8

    stats=[
        ("Total Scouts Registered",        total,                               BLUE),
        ("SVMBC Scouts",                    svmbc,                               BLUE),
        ("Out-of-Council Scouts",           ooc,                                 BLUE),
        ("T-Shirts Ordered",                tshirt_count if fmt=='A' else "N/A", BLUE),
        ("Registrations with Balance Due",  bal_count    if fmt=='A' else "N/A", RED),
        ("CampDoc Complete (100%)",         complete     if fmt=='A' else "N/A", DK_GREEN),
        ("CampDoc Not Started (0%)",        zero_cd      if fmt=='A' else "N/A", RED),
        ("Scouts w/ Discontinued Badge",    disc_count,                          RED),
    ]
    col_starts=[1,3,5,7]
    for i,(label,value,lbl_color) in enumerate(stats):
        base_row=4 if i<4 else 6; cs=col_starts[i%4]; ce=cs+1
        lc=get_column_letter(cs); lc2=get_column_letter(ce)
        ws.merge_cells(f'{lc}{base_row}:{lc2}{base_row}')
        c=ws.cell(row=base_row,column=cs,value=label)
        c.fill=fill(lbl_color); c.font=font(bold=True,color=WHITE,size=9); c.alignment=align(h="center"); ws.row_dimensions[base_row].height=18
        vr=base_row+1; ws.merge_cells(f'{lc}{vr}:{lc2}{vr}')
        c=ws.cell(row=vr,column=cs,value=value)
        c.fill=fill(LT_GRAY_BG); c.font=font(bold=True,size=16); c.alignment=align(h="center"); ws.row_dimensions[vr].height=28
    ws.row_dimensions[8].height=8

    mb_demand=Counter()
    for s in scouts:
        for mb in get_scout_mbs(s): mb_demand[mb]+=1

    mh=9; ws.merge_cells(f'A{mh}:L{mh}'); c=ws.cell(row=mh,column=1,value="Merit Badge Demand")
    c.fill=fill(DK_GREEN); c.font=font(bold=True,color=WHITE,size=11); c.alignment=align(); ws.row_dimensions[mh].height=18
    hdr(ws,mh+1,["Merit Badge","Demand","Eagle Required?","Periods","Sections Needed","Notes"],GRAY,size=9,height=16)
    dr=mh+2
    for mb,cnt in sorted(mb_demand.items(),key=lambda x:-x[1]):
        disc=is_discontinued(mb); eagle=is_eagle(mb); two_p=is_two_period(mb)
        sections="N/A" if disc else math.ceil(cnt/12)
        notes=("⚠ DISCONTINUED" if disc else ("2-period badge" if two_p else ""))
        bg=LT_RED if disc else (LT_GREEN if eagle else (WHITE if dr%2==0 else ALT_GRAY))
        for ci,val in enumerate([mb,cnt,"Yes" if eagle else "No",2 if two_p else 1,sections,notes],1):
            c=ws.cell(row=dr,column=ci,value=val); c.fill=fill(bg); c.font=font(bold=disc,size=9); c.alignment=align()
        ws.row_dimensions[dr].height=14; dr+=1

    dr+=1; ws.merge_cells(f'A{dr}:L{dr}'); c=ws.cell(row=dr,column=1,value="Enrollment by District / Council")
    c.fill=fill(BLUE); c.font=font(bold=True,color=WHITE,size=10); c.alignment=align(); ws.row_dimensions[dr].height=18; dr+=1
    hdr(ws,dr,["District","Council","Scouts","% of Total"],GRAY,size=9,height=16); dr+=1
    dist=defaultdict(lambda: defaultdict(int))
    for s in scouts: dist[s.district or '(none)'][s.council or '(none)']+=1
    svmbc_rows=sorted([(d,c,n) for d,dc in dist.items() for c,n in dc.items() if 'Silicon Valley' in c],key=lambda x:-x[2])
    ooc_rows  =sorted([(d,c,n) for d,dc in dist.items() for c,n in dc.items() if 'Silicon Valley' not in c],key=lambda x:-x[2])
    for i,(d,council,n) in enumerate(svmbc_rows+ooc_rows):
        bg=WHITE if i%2==0 else ALT_GRAY
        for ci,val in enumerate([d,council,n],1):
            c=ws.cell(row=dr,column=ci,value=val); c.fill=fill(bg); c.font=font(size=9); c.alignment=align()
        c=ws.cell(row=dr,column=4,value=n/total if total else 0)
        c.fill=fill(bg); c.font=font(size=9); c.number_format='0.0%'; ws.row_dimensions[dr].height=14; dr+=1
    ooc_total=sum(n for _,_,n in ooc_rows)
    for ci,val in enumerate(["Out-of-Council Total",'',ooc_total],1):
        c=ws.cell(row=dr,column=ci,value=val); c.fill=fill(TAN); c.font=font(bold=True,size=9)
    c=ws.cell(row=dr,column=4,value=ooc_total/total if total else 0)
    c.fill=fill(TAN); c.font=font(bold=True,size=9); c.number_format='0.0%'; ws.row_dimensions[dr].height=14
    set_col_widths(ws,[36,30,10,10,16,36]); ws.freeze_panes='A3'

# ── SCOUT ROSTER ─────────────────────────────────────────────────────────────
def build_scout_roster(wb, scouts, fmt):
    ws=wb.create_sheet("Scout Roster"); ws.sheet_properties.tabColor=MED_BLUE
    if fmt=='A':
        hdrs=["#","Scout Name","Unit","District","Council","T-Shirt","CampDoc %","Balance Due","Reg #",
              "Block 1","Block 2","Block 3","Block 4","Block 5","Block 6","Admin Comment"]
    else:
        hdrs=["#","Scout Name","Unit","District","Council","Rank","Reg #",
              "Block 1","Block 2","Block 3","Block 4","Block 5","Block 6"]
    hdr(ws,1,hdrs,BLUE,size=9,wrap=True,height=28)
    for i,s in enumerate(sorted(scouts,key=lambda s:(s.last.lower(),s.first.lower())),1):
        rn=i+1; bg=WHITE if i%2!=0 else ALT_GRAY
        mb_display=[', '.join(normalize_mb_cell(raw)) for raw in s.blocks]
        if fmt=='A':
            try: cd_val=int(s.cd_pct) if s.cd_pct else None
            except ValueError: cd_val=None
            row_vals=[i,s.name,s.unit,s.district,s.council,s.tshirt,
                      f"{cd_val}%" if cd_val is not None else 'N/A',
                      s.balance if s.balance>0 else '',s.reg_num]+mb_display+[s.admin_comment]
        else:
            row_vals=[i,s.name,s.unit,s.district,s.council,s.rank,s.reg_num]+mb_display
        for ci,val in enumerate(row_vals,1):
            c=ws.cell(row=rn,column=ci,value=val); c.fill=fill(bg); c.font=font(size=9)
            c.alignment=align(wrap=(ci>=(10 if fmt=='A' else 8)))
            if fmt=='A' and ci==8 and isinstance(val,float): c.number_format='$#,##0.00'
        ws.row_dimensions[rn].height=14
    if fmt=='A': set_col_widths(ws,[5,24,30,18,36,10,9,10,12,26,26,26,26,26,26,30])
    else:         set_col_widths(ws,[5,24,30,18,36,14,12,26,26,26,26,26,26])
    ws.freeze_panes='A2'

# ── MB SCHEDULING ────────────────────────────────────────────────────────────
def build_mb_scheduling(wb, scouts, session_label):
    ws=wb.create_sheet("MB Scheduling"); ws.sheet_properties.tabColor=DK_GREEN
    mb_demand=Counter()
    for s in scouts:
        for mb in get_scout_mbs(s): mb_demand[mb]+=1
    total=len(scouts)
    ws.merge_cells('A1:G1'); c=ws['A1']
    c.value=f"Merit Badge Scheduling  |  Camp Hi-Sierra 2026  |  {session_label}"
    c.fill=fill(DK_GREEN); c.font=font(bold=True,color=WHITE,size=13); c.alignment=align(h="center"); ws.row_dimensions[1].height=24
    ws.merge_cells('A2:G2'); c=ws['A2']
    c.value=f"1 counselor per section per week. Max 12 scouts/section. Total scouts: {total}"
    c.font=font(italic=True,size=9); c.alignment=align(); ws.row_dimensions[2].height=16
    hdr(ws,3,[f"Merit Badge",f"Enrollment\n({total})","Eagle\nRequired?","Periods",
              "Sections\nNeeded","Staff\nNeeded","Notes / Constraints"],GRAY,size=9,wrap=True,height=36)
    dr=4; aq=0
    for mb,cnt in sorted(mb_demand.items(),key=lambda x:-x[1]):
        disc=is_discontinued(mb); eagle=is_eagle(mb); two_p=is_two_period(mb)
        if disc: sections=staff="N/A"; notes="⚠ DISCONTINUED — Remove from schedule; contact all selecting scouts"
        else:
            sections=math.ceil(cnt/12)
            if mb in ('Swimming','Lifesaving'): staff="(aq. team)"; aq+=sections
            else: staff=sections
            notes="2-period badge — reserve 2 consecutive periods" if two_p else ""
        bg=LT_RED if disc else (LT_GREEN if eagle else (WHITE if dr%2==0 else ALT_GRAY))
        for ci,val in enumerate([mb,cnt,"Yes" if eagle else "No",2 if two_p else 1,sections,staff,notes],1):
            c=ws.cell(row=dr,column=ci,value=val)
            c.fill=fill(bg); c.font=font(bold=disc,size=9); c.alignment=align(wrap=(ci==7))
        ws.row_dimensions[dr].height=14; dr+=1
    ws.merge_cells(f'A{dr}:G{dr}')
    c=ws.cell(row=dr,column=1,value=f"Aquatics team: {aq} staff-sections shared across Swimming and Lifesaving.")
    c.fill=fill(TAN); c.font=font(italic=True,size=9); ws.row_dimensions[dr].height=14
    set_col_widths(ws,[36,12,10,8,12,10,48]); ws.freeze_panes='A4'

# ── DISCONTINUED BADGE OUTREACH ───────────────────────────────────────────────
def build_discontinued(wb, scouts):
    ws=wb.create_sheet("Discontinued Badge Outreach"); ws.sheet_properties.tabColor=RED
    ws.merge_cells('A1:J1'); c=ws['A1']; c.value="Discontinued Badge Outreach  |  Camp Hi-Sierra 2026"
    c.fill=fill(RED); c.font=font(bold=True,color=WHITE,size=13); c.alignment=align(h="center"); ws.row_dimensions[1].height=22
    hdr(ws,2,["#","Scout Name","Unit","District","Discontinued Badge","Block Slot",
              "Parent/Scout Email","Parent Phone","Contacted?","Notes"],GRAY,size=9,wrap=True,height=28)
    disc_rows=[]
    for s in scouts:
        for slot_idx,raw in enumerate(s.blocks,1):
            for mb in normalize_mb_cell(raw):
                if is_discontinued(mb): disc_rows.append((s,f"Block {slot_idx}",mb))
    if not disc_rows:
        ws.merge_cells('A3:J3'); c=ws['A3']
        c.value="✓ No scouts have selected discontinued badges."
        c.fill=fill(LT_GREEN); c.font=font(italic=True,size=10,color=DK_GREEN); c.alignment=align(h="center")
    else:
        for i,(s,slot,mb) in enumerate(disc_rows,1):
            rn=i+2; bg=LT_RED if i%2!=0 else "FCEAEA"
            for ci,val in enumerate([i,s.name,s.unit,s.district,mb,slot,s.parent_email,s.parent_phone,'',''],1):
                c=ws.cell(row=rn,column=ci,value=val); c.fill=fill(bg); c.font=font(size=9); c.alignment=align()
            ws.row_dimensions[rn].height=14
    set_col_widths(ws,[5,24,30,18,36,8,32,16,10,30]); ws.freeze_panes='A3'

# ── BALANCE DUE ──────────────────────────────────────────────────────────────
def build_balance_due(wb, scouts, fmt):
    ws=wb.create_sheet("Balance Due"); ws.sheet_properties.tabColor=ORANGE
    ws.merge_cells('A1:I1'); c=ws['A1']; c.value="Balance Due  |  Camp Hi-Sierra 2026"
    c.fill=fill(ORANGE); c.font=font(bold=True,color=WHITE,size=13); c.alignment=align(h="center"); ws.row_dimensions[1].height=22
    if fmt=='B':
        ws.merge_cells('A2:I2'); c=ws['A2']
        c.value="Balance information not available in Class Schedule Export format."
        c.font=font(italic=True,size=10); c.alignment=align(h="center")
        set_col_widths(ws,[5,24,30,18,14,12,32,16,30]); return
    hdr(ws,2,["#","Scout Name","Unit","District","Balance Due ($)","Reg #",
              "Parent Email","Parent Phone","Admin Comment"],GRAY,size=9,wrap=True,height=28)
    with_bal=sorted([s for s in scouts if s.balance>0],key=lambda s:-s.balance)
    for i,s in enumerate(with_bal,1):
        rn=i+2; bg=WHITE if i%2!=0 else ALT_GRAY
        for ci,val in enumerate([i,s.name,s.unit,s.district,s.balance,s.reg_num,s.parent_email,s.parent_phone,s.admin_comment],1):
            c=ws.cell(row=rn,column=ci,value=val); c.fill=fill(bg); c.font=font(size=9); c.alignment=align()
            if ci==5: c.number_format='$#,##0.00'
        ws.row_dimensions[rn].height=14
    tr=len(with_bal)+3; total_bal=sum(s.balance for s in with_bal)
    ws.cell(row=tr,column=1,value="TOTAL").fill=fill(ORANGE); ws.cell(row=tr,column=1).font=font(bold=True,color=WHITE,size=10)
    c=ws.cell(row=tr,column=5,value=total_bal); c.fill=fill(ORANGE); c.font=font(bold=True,color=WHITE,size=10)
    c.number_format='$#,##0.00'; ws.row_dimensions[tr].height=18
    set_col_widths(ws,[5,24,30,18,14,12,32,16,30]); ws.freeze_panes='A3'

# ── CAMPDOC STATUS ────────────────────────────────────────────────────────────
def build_campdoc(wb, scouts, fmt):
    ws=wb.create_sheet("CampDoc Status"); ws.sheet_properties.tabColor=PURPLE
    ws.merge_cells('A1:H1'); c=ws['A1']; c.value="CampDoc Status  |  Camp Hi-Sierra 2026"
    c.fill=fill(PURPLE); c.font=font(bold=True,color=WHITE,size=13); c.alignment=align(h="center"); ws.row_dimensions[1].height=22
    if fmt=='B':
        ws.merge_cells('A2:H2'); c=ws['A2']
        c.value="CampDoc information not available in Class Schedule Export format."
        c.font=font(italic=True,size=10); c.alignment=align(h="center")
        set_col_widths(ws,[5,24,30,18,32,10,13,12]); return
    hdr(ws,2,["#","Scout Name","Unit","District","CampDoc Email","CampDoc %","Balance Due ($)","Reg #"],
        GRAY,size=9,wrap=True,height=28)
    def sk(s):
        if not s.cd_email: return (3,0)
        try: pct=int(s.cd_pct or '0')
        except ValueError: pct=0
        return (0,0) if pct==0 else ((2,100) if pct==100 else (1,pct))
    for i,s in enumerate(sorted(scouts,key=sk),1):
        rn=i+2
        try: cd_pct=int(s.cd_pct or '0')
        except ValueError: cd_pct=0
        bg=(LT_GRAY if not s.cd_email else LT_RED if cd_pct==0 else LT_GREEN if cd_pct==100 else LT_YELLOW)
        for ci,val in enumerate([i,s.name,s.unit,s.district,s.cd_email,
                                  f"{cd_pct}%" if s.cd_email else 'No CampDoc',
                                  s.balance if s.balance>0 else '',s.reg_num],1):
            c=ws.cell(row=rn,column=ci,value=val); c.fill=fill(bg); c.font=font(size=9); c.alignment=align()
            if ci==7 and isinstance(val,float): c.number_format='$#,##0.00'
        ws.row_dimensions[rn].height=14
    set_col_widths(ws,[5,24,30,18,32,10,13,12]); ws.freeze_panes='A3'

# ── CSV LOADER ────────────────────────────────────────────────────────────────
def load_csv(csv_path):
    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader=csv.DictReader(f); headers=reader.fieldnames; raw_rows=list(reader)
    fmt=detect_format(headers)
    scouts=parse_fmt_a(raw_rows) if fmt=='A' else parse_fmt_b(raw_rows)
    sessions=sorted({r.get('Session','') for r in raw_rows if r.get('Session','')})
    return fmt, scouts, sessions

# ── FIRESTORE PUSH ────────────────────────────────────────────────────────────
def session_doc_id(session_name):
    """'Week 3' → 'Week_3',  'Eagle Achievement Camp 2026' → 'Eagle_Achievement_Camp_2026'"""
    return re.sub(r'\s+', '_', session_name.strip())

def load_class_sizes(csv_path):
    """
    Parse Class_Attendee_Counts CSV → {normalized_badge: total_capacity_across_blocks}
    Also returns granular {normalized_badge: {block: capacity}} for per-block detail.
    """
    by_badge_block = {}  # {normalized_badge: {block: max_capacity}}
    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_name  = row.get('Class Name', '').strip()
            block     = row.get('Period', '').strip()
            max_size  = row.get('Max Class Size', '').strip()
            if not raw_name or not block or not max_size:
                continue
            try:
                max_size = int(max_size)
            except ValueError:
                continue
            # Normalize using the same function as scout block data
            badges = normalize_mb_cell(raw_name)
            for badge in badges:
                if badge not in by_badge_block:
                    by_badge_block[badge] = {}
                # Sum capacity when the same badge/block appears more than once (multiple sections)
                by_badge_block[badge][block] = by_badge_block[badge].get(block, 0) + max_size

    # Aggregate total capacity per badge across all blocks
    total_capacity = {badge: sum(blocks.values()) for badge, blocks in by_badge_block.items()}
    return total_capacity, by_badge_block


def push_to_firestore(scouts, fmt, service_account_path, class_sizes_path=None):
    """Push mbDemand + session stats to Firestore sessions/{sessionId} for each week."""
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore as fs
    except ImportError:
        print("⚠  firebase-admin not installed. Run: pip install firebase-admin --break-system-packages")
        return

    # Init (or reuse existing) Firebase app
    try:
        app = firebase_admin.get_app()
    except ValueError:
        cred = credentials.Certificate(service_account_path)
        app = firebase_admin.initialize_app(cred)

    db = fs.client()

    # Load class capacities if provided
    total_capacity, block_capacity = {}, {}
    if class_sizes_path:
        total_capacity, block_capacity = load_class_sizes(class_sizes_path)
        print(f"  Loaded capacity for {len(total_capacity)} badge(s) from class sizes CSV.")

    # Group scouts by their Session field
    by_session = defaultdict(list)
    for s in scouts:
        sess = s.session or 'Unknown'
        by_session[sess].append(s)

    for sess_name, sess_scouts in sorted(by_session.items()):
        doc_id = session_doc_id(sess_name)
        total  = len(sess_scouts)
        svmbc  = sum(1 for s in sess_scouts if 'Silicon Valley' in s.council)

        # mbDemand: normalized badge name → count (per-scout deduplicated)
        mb_demand = Counter()
        for s in sess_scouts:
            for mb in get_scout_mbs(s): mb_demand[mb] += 1

        # Compute spotsRemaining per badge using class sizes CSV
        spots_remaining = {}
        if total_capacity:
            all_badges = set(mb_demand.keys()) | set(total_capacity.keys())
            for badge in all_badges:
                cap = total_capacity.get(badge, 0)
                dem = mb_demand.get(badge, 0)
                spots_remaining[badge] = max(0, cap - dem)

        # CampDoc & balance (fmt A only)
        if fmt == 'A':
            cd_complete = sum(1 for s in sess_scouts if s.cd_pct == '100')
            cd_started  = sum(1 for s in sess_scouts if s.cd_email and s.cd_pct not in ('','0'))
            cd_total    = sum(1 for s in sess_scouts if s.cd_email)
            cd_nums     = [int(s.cd_pct) for s in sess_scouts if s.cd_pct.isdigit() and s.cd_email]
            avg_cd      = round(sum(cd_nums)/len(cd_nums), 1) if cd_nums else 0
            balance_due = round(sum(s.balance for s in sess_scouts if s.balance > 0), 2)
        else:
            cd_complete=cd_started=cd_total=avg_cd=balance_due=0

        units_set = sorted({s.unit for s in sess_scouts if s.unit})

        doc_data = {
            'sessionName':          sess_name,
            'type':                 'summer_camp',
            'youthCount':           total,
            'totalRegistrants':     total,
            'totalCount':           total,
            'adultCount':           0,          # youth-only in this export
            'unitCount':            len(units_set),
            'svmbcCount':           svmbc,
            'outOfCouncilCount':    total - svmbc,
            'mbDemand':             dict(mb_demand),
            'mbCapacity':           total_capacity if total_capacity else {},
            'mbSpotsRemaining':     spots_remaining if spots_remaining else {},
            'balanceDue':           balance_due,
            'totalBalanceDue':      balance_due,
            'campdocTotalCount':    cd_total,
            'campdocStartedCount':  cd_started,
            'campdocCompleteCount': cd_complete,
            'avgCampdocPct':        avg_cd,
            'lastSync':             int(datetime.utcnow().timestamp() * 1000),
            'lastSyncSource':       ('BlackPug_EventDataDump' if fmt == 'A' else 'myEventClassData'),
        }

        print(f"  [{doc_id}] Writing session metadata...", end=' ', flush=True)
        try:
            db.collection('sessions').document(doc_id).set(doc_data, merge=True)
            print("✅")
        except Exception as e:
            print(f"❌ Error: {e}")
            continue

        # Also write mbDemand to the subcollection the app reads for "Last Merit Badge update"
        if mb_demand:
            print(f"  [{doc_id}] Writing mbDemand subcollection...", end=' ', flush=True)
            try:
                mb_doc = {
                    'demand':         dict(mb_demand),
                    'updatedAt':      fs.SERVER_TIMESTAMP,
                }
                if spots_remaining:
                    mb_doc['capacity']       = total_capacity
                    mb_doc['spotsRemaining'] = spots_remaining
                db.collection('sessions').document(doc_id)\
                  .collection('mbDemand').document('current')\
                  .set(mb_doc, merge=True)
                print("✅")
            except Exception as e:
                print(f"❌ Error: {e}")

    print(f"Firestore push complete — {len(by_session)} session(s) updated.")

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: build_chs_workbook.py <file.csv> [file2.csv ...] [output_dir]")
        print("       Add --service-account <key.json> to also push to Firestore")
        print("       Add --push-only <key.json> to skip Excel and only push to Firestore")
        print("       Add --class-sizes <counts.csv> to include spotsRemaining in push")
        sys.exit(1)

    args=sys.argv[1:]
    out_dir="."; service_account=None; push_only=False; class_sizes=None

    if '--service-account' in args:
        idx=args.index('--service-account')
        service_account=args[idx+1]; args=args[:idx]+args[idx+2:]
    if '--push-only' in args:
        idx=args.index('--push-only')
        service_account=args[idx+1]; args=args[:idx]+args[idx+2:]; push_only=True
    if '--class-sizes' in args:
        idx=args.index('--class-sizes')
        class_sizes=args[idx+1]; args=args[:idx]+args[idx+2:]
    if len(args)>1 and not args[-1].lower().endswith('.csv'):
        out_dir=args[-1]; args=args[:-1]

    all_scouts=[]; all_sessions=[]; fmt=None
    for csv_path in args:
        print(f"Loading {os.path.basename(csv_path)}...")
        f,scouts,sessions=load_csv(csv_path)
        if fmt is None: fmt=f
        elif fmt!=f: print(f"  ⚠ Mixed format — merging A+B as B (no balance/CampDoc)"); fmt='B'
        print(f"  {'Event_Data_Dump' if f=='A' else 'myEventClassData'} | {len(scouts)} scouts | sessions: {', '.join(sessions)}")
        all_scouts.extend(scouts); all_sessions.extend(sessions)

    scouts=all_scouts; effective_fmt=fmt or 'B'
    unique_sessions=sorted(set(all_sessions))
    session_label=', '.join(unique_sessions) if unique_sessions else "Camp 2026"
    print(f"\nTotal scouts: {len(scouts)}")

    # ── Firestore push ──
    if service_account:
        print(f"\nPushing to Firestore ({len(set(s.session for s in scouts))} session(s))...")
        push_to_firestore(scouts, effective_fmt, service_account, class_sizes)

    if push_only:
        _print_summary(scouts, effective_fmt)
        return

    # ── Excel workbook ──
    wb=Workbook(); wb.remove(wb.active)
    print("\nBuilding Dashboard..."); build_dashboard(wb,scouts,session_label,effective_fmt)
    print("Building Scout Roster..."); build_scout_roster(wb,scouts,effective_fmt)
    print("Building MB Scheduling..."); build_mb_scheduling(wb,scouts,session_label)
    print("Building Discontinued Badge Outreach..."); build_discontinued(wb,scouts)
    print("Building Balance Due..."); build_balance_due(wb,scouts,effective_fmt)
    print("Building CampDoc Status..."); build_campdoc(wb,scouts,effective_fmt)

    today=date.today().strftime("%Y%m%d")
    slug=re.sub(r'Week\s+','Wk',session_label).replace(', ','-').replace(' ','')
    out_path=os.path.join(out_dir,f"CHS_2026_{slug}_Scheduling_{today}.xlsx")
    wb.save(out_path); print(f"\n✅ Saved: {out_path}")
    _print_summary(scouts, effective_fmt)

def _print_summary(scouts, effective_fmt):
    print("\n── POST-OUTPUT SUMMARY ─────────────────────────────")
    print(f"Total scouts: {len(scouts)}")
    disc_hits=[(s.full_name,mb) for s in scouts for mb in get_scout_mbs(s) if is_discontinued(mb)]
    print(f"Discontinued badge selections: {len(disc_hits)}"+(" ✓" if not disc_hits else ""))
    for name,mb in disc_hits[:10]: print(f"  {name}: {mb}")
    if effective_fmt=='A':
        with_bal=[s for s in scouts if s.balance>0]; total_bal=sum(s.balance for s in with_bal)
        print(f"Balance due: {len(with_bal)} scouts, ${total_bal:,.2f} total")
        complete=sum(1 for s in scouts if s.cd_pct=='100')
        zero_cd =sum(1 for s in scouts if s.cd_pct=='0' and s.cd_email)
        partial =sum(1 for s in scouts if s.cd_email and s.cd_pct not in ('0','100'))
        no_email=sum(1 for s in scouts if not s.cd_email)
        print(f"CampDoc: 100%={complete}, partial={partial}, 0%={zero_cd}, no email={no_email}")
    else:
        print("Balance due / CampDoc: N/A (class schedule format)")
    names=[s.full_name for s in scouts]
    dupes=[n for n,c in Counter(names).items() if c>1]
    print(f"Duplicate names: {dupes if dupes else 'None ✓'}")

if __name__ == '__main__':
    main()