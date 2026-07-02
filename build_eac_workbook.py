#!/usr/bin/env python3
"""
EAC (Eagle Achievement Camp) workbook builder
Produces the 10-tab EAC_YYYY_Scheduling_YYYYMMDD.xlsx from an Event_Data_Dump CSV.
"""
import csv, re, sys, os
from collections import defaultdict, Counter
from datetime import date, datetime
from math import ceil
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as num_fmt)
from openpyxl.utils import get_column_letter

# ── BRAND COLORS ──────────────────────────────────────────────────────────────
BLUE    ="003F87"; MED_BLUE="2E75B6"; DK_GREEN="1F7A4D"
RED     ="CE1126"; ORANGE  ="ED7D31"; PURPLE  ="7030A0"
GRAY    ="515354"; TAN     ="D6CEBD"
WHITE   ="FFFFFF"; LT_GRAY ="F2F2F2"; ALT_GRAY="F2F2F2"
LT_GREEN="E2EFDA"; LT_YELLOW="FFF2CC"; LT_RED  ="FCE4D6"
LT_BLUE ="D0E4FF"; LT_BLUE2="BDD7EE"; TEAL    ="EBFFFE"

FILL   = lambda c: PatternFill("solid", fgColor=c)
FONT   = lambda bold=False, color=WHITE, size=10, italic=False: Font(
            name="Arial", bold=bold, color=color, size=size, italic=italic)
ALIGN  = lambda h="left", v="center", wrap=False: Alignment(
            horizontal=h, vertical=v, wrap_text=wrap)
BORDER = lambda: Border(
    left  =Side(style='thin',color='DDDDDD'),
    right =Side(style='thin',color='DDDDDD'),
    top   =Side(style='thin',color='DDDDDD'),
    bottom=Side(style='thin',color='DDDDDD'))

# ── YOUTH TYPES ────────────────────────────────────────────────────────────────
YOUTH_KEYS = ('Youth','CIT','BSA Youth')   # substring match

# ── EAGLE REQUIRED / 2-PERIOD ─────────────────────────────────────────────────
EAGLE_REQUIRED = {
    "First Aid","Citizenship in the Community","Citizenship in the Nation",
    "Citizenship in the World","Communications","Communication","Cooking",
    "Personal Fitness","Emergency Preparedness","Lifesaving",
    "Environmental Science","Sustainability","Personal Management",
    "Swimming","Cycling","Hiking & Backpacking","Camping","Family Life",
}
TWO_PERIOD = {"Archery","Climbing","Rifle Shooting","Shotgun Shooting","Welding"}

def normalize_mb(raw):
    """Normalize a raw block/preference value into a list of clean badge names."""
    if not raw or not raw.strip(): return []
    raw = raw.strip()
    if any(p in raw.lower() for p in ['(no badge)', 'spl 101', 'n/a']): return []
    # Split combined entries like "Backpacking (2024), Hiking (2023)"
    parts = re.split(r',\s+(?=[A-Z])', raw)
    result = []
    for part in parts:
        b = part.strip()
        b = re.sub(r'\s*\(\d{4}\s+version\)', '', b)          # strip (YYYY version)
        b = re.sub(r'\s*\(\d+\s+[Ss]ession\)', '', b)          # strip (N Session)
        b = re.sub(r'\s*[-]\s*(Morning|Afternoon)\s*$', '', b, flags=re.I)
        if re.search(r'citizenship in (the )?society', b, re.I):
            b = 'Citizenship in Society (DISCONTINUED)'
        if re.search(r'backpacking', b, re.I) and re.search(r'hiking', b, re.I):
            b = 'Hiking & Backpacking'
        b = b.replace('Substainablity','Sustainability')
        b = b.replace('Communication ','Communications ').strip()
        if b.lower() == 'communication': b = 'Communications'
        b = b.strip()
        if b:
            result.append(b)
    return result

def is_eagle(mb):       return mb in EAGLE_REQUIRED or mb.replace('Communication','Communications') in EAGLE_REQUIRED
def is_two_period(mb):  return any(tp.lower() in mb.lower() for tp in TWO_PERIOD)
def is_discontinued(mb):return 'DISCONTINUED' in mb

# ── SCOUT DATACLASS ───────────────────────────────────────────────────────────
class Scout:
    __slots__ = ('first','last','unit','district','council','reg_num','rank',
                 'tshirt','bus','balance','cd_email','cd_pct','parent_email',
                 'parent_phone','parent_name','admin_comment',
                 'blocks','pref_slots','registrant_type','session','event_name')
    def __init__(self,**kw):
        for k in self.__slots__: setattr(self, k, kw.get(k,''))
        if not self.blocks: self.blocks = ['']*5
        if not self.pref_slots: self.pref_slots = ['']*6
    @property
    def name(self):      return f"{self.last}, {self.first}"
    @property
    def full_name(self): return f"{self.first} {self.last}"
    @property
    def block_mbs(self):
        """Normalized badge names from block assignments (deduplicated)."""
        seen = set(); result = []
        for raw in self.blocks:
            for mb in normalize_mb(raw):
                if mb not in seen:
                    seen.add(mb); result.append(mb)
        return result
    @property
    def pref_mbs(self):
        """Normalized badge names from free-text preferences (deduplicated)."""
        seen = set(); result = []
        for raw in self.pref_slots:
            for mb in normalize_mb(raw):
                if mb not in seen:
                    seen.add(mb); result.append(mb)
        return result

# ── CSV PARSING ───────────────────────────────────────────────────────────────
PREF_COLS = [
    'x1st Merit Badge Class Selection',
    'x2nd Merit Badge Class Selection',
    'x3rd Merit Badge Class Selection',
    'x4th Merit Badge Class Selection',
    'x5th Merit Badge Selection',
    'x6th Merit Badge Choice',
]

def load_csv(path):
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    event_name = next((r.get('Event','') for r in rows if r.get('Event','')), '2026 Eagle Achievement Camp')
    session_start = next((r.get('Session Start Date','') for r in rows if r.get('Session Start Date','')), '')

    scouts = []
    unnamed_regs = []
    for r in rows:
        fname = r.get('First Name','').strip()
        rtype = r.get('Registrant Type','').strip()
        if not any(k in rtype for k in YOUTH_KEYS): continue
        if not fname:
            unnamed_regs.append(r)
            continue
        if r.get('Cancelled','').strip().lower() in ('yes','1','true'): continue
        try: balance = float(r.get('Balance Due') or 0)
        except ValueError: balance = 0.0
        # T-shirt
        tshirt = ''
        for sz in ['Youth Small','Youth Medium','Youth Large','Youth X-Large',
                   'Adult Small','Adult Medium','Adult Large','Adult X-Large','Adult 2X-Large']:
            if r.get(f'T-Shirt Size - {sz}','') == '1':
                tshirt = sz; break
        bus = '1' if r.get('Provided Bus Transportation - YES I will ride the bus','') == '1' else ''
        blocks = [normalize_mb_raw(r.get(f'Block {i}','')) for i in range(1,6)]
        pref_slots = [r.get(c,'').strip() for c in PREF_COLS]
        scouts.append(Scout(
            first=fname, last=r.get('Last Name','').strip(),
            unit=r.get('Unit','').strip(), district=r.get('District','').strip(),
            council=r.get('Council','').strip(), reg_num=r.get('Registration Number','').strip(),
            rank=r.get('Scouting Rank','').strip(), tshirt=tshirt, bus=bus,
            balance=balance, cd_email=r.get('CampDoc Email','').strip(),
            cd_pct=r.get('CampDoc Completeness','').strip(),
            parent_email=r.get('Parent Guardian Email','').strip(),
            parent_phone=r.get('Parent Guardian Phone','').strip(),
            parent_name=r.get('Parent Guardian Name','').strip(),
            admin_comment=r.get('Admin Comment','').strip(),
            blocks=blocks, pref_slots=pref_slots,
            registrant_type=rtype,
            session=r.get('Session','').strip(),
            event_name=event_name,
        ))
    return scouts, event_name, session_start, unnamed_regs

def normalize_mb_raw(raw):
    """Return the raw string unmodified (stored; normalized lazily)."""
    return raw.strip() if raw else ''

# ── SECTION ASSIGNMENT ────────────────────────────────────────────────────────
def build_section_assignments(scouts):
    """
    Returns: {(badge, block_idx): [(scout, section_num), ...]}
    block_idx is 0-based (0=Block1…4=Block5).
    Scouts in the same (badge, block) are sorted by unit then name and capped at 12/section.
    """
    groups = defaultdict(list)
    for s in scouts:
        for bi, raw in enumerate(s.blocks):
            for mb in normalize_mb(raw):
                groups[(mb, bi)].append(s)

    result = {}
    for key, group in groups.items():
        group_sorted = sorted(group, key=lambda s: (s.unit, s.name))
        sections = []
        for i, s in enumerate(group_sorted):
            sec_num = i // 12 + 1
            sections.append((s, sec_num))
        result[key] = sections
    return result

def scout_schedule(scouts, section_assignments):
    """
    Returns per-scout schedule: {scout: [(badge_or_None, section_or_None), ...] for blocks 0-4}
    """
    # Invert section_assignments to {(scout_id, block): (badge, sec)}
    lookup = {}
    for (mb, bi), entries in section_assignments.items():
        for (s, sec_num) in entries:
            lookup[(id(s), bi)] = (mb, sec_num)
    result = {}
    for s in scouts:
        sched = []
        for bi in range(5):
            if (id(s), bi) in lookup:
                sched.append(lookup[(id(s), bi)])
            else:
                sched.append((None, None))
        result[id(s)] = sched
    return result

# ── OPENPYXL HELPERS ──────────────────────────────────────────────────────────
def hdr(ws, row, vals, bg, fg=WHITE, bold=True, size=10, height=18, wrap=False):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.fill = FILL(bg); c.font = FONT(bold=bold, color=fg, size=size)
        c.alignment = ALIGN(wrap=wrap)
    ws.row_dimensions[row].height = height

def set_cell(ws, row, col, val, bg=None, fg='000000', bold=False, size=10,
             halign='left', wrap=False, italic=False, border=True, num_format=None):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = FILL(bg)
    c.font = Font(name="Arial", bold=bold, color=fg, size=size, italic=italic)
    c.alignment = ALIGN(h=halign, wrap=wrap)
    if border: c.border = BORDER()
    if num_format: c.number_format = num_format
    return c

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── MB DEMAND (from preferences) ─────────────────────────────────────────────
def pref_mb_demand(scouts):
    counter = Counter()
    for s in scouts:
        for mb in s.pref_mbs:
            counter[mb] += 1
    return counter

def block_mb_demand(scouts):
    counter = Counter()
    for s in scouts:
        for mb in s.block_mbs:
            counter[mb] += 1
    return counter

# ── TAB 1: DASHBOARD ──────────────────────────────────────────────────────────
def build_dashboard(wb, scouts, event_name, session_start, fmt_note=''):
    ws = wb.create_sheet("Dashboard"); ws.sheet_properties.tabColor = BLUE
    TARGET = 175

    total    = len(scouts)
    svmbc    = sum(1 for s in scouts if 'Silicon Valley' in s.council)
    ooc      = total - svmbc
    bus_cnt  = sum(1 for s in scouts if s.bus == '1')
    bal_cnt  = sum(1 for s in scouts if s.balance > 0)
    cd_comp  = sum(1 for s in scouts if s.cd_pct == '100')
    cd_zero  = sum(1 for s in scouts if s.cd_pct == '0' and s.cd_email)
    disc_cnt = sum(1 for s in scouts if any(is_discontinued(mb) for mb in s.pref_mbs))
    today_str = date.today().strftime("%B %d, %Y")

    # Title row
    ws.merge_cells('A1:N1')
    c = ws['A1']; c.value = f"Camp Hi-Sierra  |  Eagle Achievement Camp 2026"
    c.fill=FILL(BLUE); c.font=FONT(bold=True, size=16); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:N2')
    c = ws['A2']; c.value = f"Camp Hi-Sierra  |  {session_start}  |  Generated {today_str}"
    c.fill=FILL(GRAY); c.font=FONT(size=11); c.alignment=ALIGN(h='center')
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    # Stats block 8 stats (2-col merged each)
    stats = [
        ("Total Scouts Registered",      total,     BLUE),
        ("SVMBC Scouts",                 svmbc,     BLUE),
        ("Out-of-Council Scouts",        ooc,       BLUE),
        ("Bus Riders",                   bus_cnt,   BLUE),
        ("Registrations with Balance Due",bal_cnt,  RED),
        ("CampDoc Complete (100%)",      cd_comp,   DK_GREEN),
        ("CampDoc Not Started (0%)",     cd_zero,   RED),
        ("Scouts Selecting Discontinued",disc_cnt,  RED),
    ]
    stat_row_lbl = 4; stat_row_val = 5
    ws.row_dimensions[stat_row_lbl].height = 16
    ws.row_dimensions[stat_row_val].height = 40
    ws.row_dimensions[6].height = 8

    for i, (label, value, color) in enumerate(stats):
        col = i * 2 + 1
        ws.merge_cells(start_row=stat_row_lbl, start_column=col, end_row=stat_row_lbl, end_column=col+1)
        ws.merge_cells(start_row=stat_row_val, start_column=col, end_row=stat_row_val, end_column=col+1)
        lc = ws.cell(row=stat_row_lbl, column=col, value=label)
        lc.fill=FILL(color); lc.font=FONT(color=WHITE, size=9)
        lc.alignment=ALIGN(h='center')
        vc = ws.cell(row=stat_row_val, column=col, value=value)
        vc.fill=FILL(LT_GRAY); vc.font=Font(name="Arial", bold=True, size=16, color='000000')
        vc.alignment=ALIGN(h='center')

    # EAC projection note
    ws.merge_cells('A7:N7')
    nc = ws['A7']
    nc.value = f"Note: Sections are calculated for {TARGET}-scout target enrollment at 12 scouts/section."
    nc.fill=FILL(DK_GREEN); nc.font=FONT(italic=True, size=10)
    nc.alignment=ALIGN(h='center')
    ws.row_dimensions[7].height = 16

    # MB demand header
    ws.merge_cells('A8:N8')
    mc = ws['A8']
    mc.value = f"Merit Badge Demand  |  Sections Scaled to {TARGET} Scouts"
    mc.fill=FILL(DK_GREEN); mc.font=FONT(bold=True, size=11)
    mc.alignment=ALIGN(h='center')
    ws.row_dimensions[8].height = 18

    mb_cols = ["Merit Badge","Current Enrollment","Eagle Required?","Periods",
               f"Sections @ {TARGET} scouts (12/sect)","Staff Needed","Notes"]
    hdr(ws, 9, mb_cols, GRAY, size=9, height=16, wrap=True)
    ws.merge_cells('A9:C9') if False else None  # no merges on header

    demand = pref_mb_demand(scouts)
    mb_row = 10
    for mb, cnt in sorted(demand.items(), key=lambda x: -x[1]):
        disc = is_discontinued(mb)
        eagle = is_eagle(mb)
        two_p = is_two_period(mb)
        periods = 2 if two_p else 1
        if disc:
            sections_175 = 'N/A'; staff = 'N/A'; notes = "⚠ DISCONTINUED — do not schedule"
            bg = LT_RED
        else:
            sections_175 = ceil(TARGET * (cnt/total) / 12) if total else 0
            staff = sections_175 if not disc else 'N/A'
            notes = "2-period badge (needs consecutive blocks)" if two_p else ("Aquatics team" if mb in ("Swimming","Lifesaving") else "")
            bg = LT_GREEN if eagle else (WHITE if (mb_row-10)%2==0 else ALT_GRAY)
        for ci, val in enumerate([mb, cnt, "Yes" if eagle else "No", periods,
                                  sections_175, staff, notes], 1):
            set_cell(ws, mb_row, ci, val, bg=bg,
                     fg='000000' if not disc else '8B0000',
                     bold=disc, halign='center' if ci>1 else 'left')
        mb_row += 1

    # District enrollment
    mb_row += 1
    ws.merge_cells(start_row=mb_row, start_column=1, end_row=mb_row, end_column=7)
    dc = ws.cell(row=mb_row, column=1, value="Enrollment by District / Council")
    dc.fill=FILL(BLUE); dc.font=FONT(bold=True, size=11); dc.alignment=ALIGN(h='center')
    ws.row_dimensions[mb_row].height = 18; mb_row += 1

    hdr(ws, mb_row, ["District","Council","# Scouts","% of Total"], GRAY, height=16)
    mb_row += 1

    by_dist = Counter()
    by_dist_council = {}
    for s in scouts:
        key = (s.district or '(No District)', s.council or '(No Council)')
        by_dist[key] += 1
        by_dist_council[key] = s.council
    ooc_total = 0
    for di, (key, cnt) in enumerate(sorted(by_dist.items(), key=lambda x: -x[1])):
        dist, council = key
        is_ooc = 'Silicon Valley' not in council
        if is_ooc: ooc_total += cnt
        bg = WHITE if di%2==0 else ALT_GRAY
        pct_formula = f"=C{mb_row}/{total}"
        for ci, val in enumerate([dist, council, cnt, pct_formula], 1):
            sc = set_cell(ws, mb_row, ci, val, bg=bg, halign='center' if ci>2 else 'left')
            if ci == 4:
                sc.number_format = '0.0%'
        mb_row += 1

    for ci, val in enumerate(["Out-of-Council Total","",ooc_total,f"={ooc_total}/{total}"],1):
        sc = set_cell(ws, mb_row, ci, val, bg=TAN, bold=True, halign='center' if ci>2 else 'left')
        if ci == 4:
            sc.number_format = '0.0%'

    col_widths(ws, [36,22,14,10,30,14,36])
    ws.freeze_panes = 'A10'

# ── TAB 2: SCOUT ROSTER ───────────────────────────────────────────────────────
def build_scout_roster(wb, scouts):
    ws = wb.create_sheet("Scout Roster"); ws.sheet_properties.tabColor = MED_BLUE
    ws.merge_cells('A1:N1')
    c = ws['A1']; c.value = "Eagle Achievement Camp 2026  |  Scout Roster"
    c.fill=FILL(MED_BLUE); c.font=FONT(bold=True,size=14); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 24
    hdr(ws, 2, ["#","Scout Name","Rank","Unit","District","Council","Bus","T-Shirt",
                "CampDoc %","Balance Due","Reg #","Parent Name","Parent Email","Parent Phone"],
        GRAY, size=9, height=16, wrap=True)
    sorted_scouts = sorted(scouts, key=lambda s: s.name)
    for i, s in enumerate(sorted_scouts, 1):
        bg = WHITE if i%2==0 else LT_GRAY
        cd_pct_val = (int(s.cd_pct) if s.cd_pct.isdigit() else '') if s.cd_email else ''
        row = i+2
        for ci, val in enumerate([i, s.full_name, s.rank, s.unit, s.district, s.council,
                                   'Yes' if s.bus=='1' else '', s.tshirt,
                                   cd_pct_val, s.balance if s.balance else '',
                                   s.reg_num, s.parent_name, s.parent_email, s.parent_phone], 1):
            sc = set_cell(ws, row, ci, val, bg=bg, halign='center' if ci in (1,7,8,9,10) else 'left')
            if ci == 10 and s.balance > 0:
                sc.number_format = '"$"#,##0.00'
    col_widths(ws, [5,26,12,30,18,22,6,14,10,12,13,20,28,16])
    ws.freeze_panes = 'A3'

# ── TAB 3: MB SCHEDULING ─────────────────────────────────────────────────────
def build_mb_scheduling(wb, scouts, event_name):
    ws = wb.create_sheet("MB Scheduling"); ws.sheet_properties.tabColor = DK_GREEN
    TARGET = 175
    total = len(scouts)

    ws.merge_cells('A1:J1')
    c = ws['A1']; c.value = f"Eagle Achievement Camp 2026  |  Merit Badge Scheduling"
    c.fill=FILL(DK_GREEN); c.font=FONT(bold=True,size=14); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:J2')
    nc = ws['A2']
    nc.value = "Staffing model: 1 counselor = 5 block-slots (Block 1 through Block 5). Max 12 scouts/section."
    nc.font = Font(name="Arial", italic=True, size=9, color='000000')
    nc.alignment = ALIGN(h='left')
    ws.row_dimensions[2].height = 14

    hdrs = [f"Merit Badge",f"Enrollment\n({total})","Eagle\nRequired?","Periods",
            f"Sections\n@ {total} scouts",f"Sections\n@ {TARGET} scouts",
            "Staff per\nSection","Staff Needed\n@ actual","Staff Needed\n@ 175","Notes / Constraints"]
    hdr(ws, 3, hdrs, GRAY, size=9, height=36, wrap=True)

    # Color bands for actual vs 175 columns
    for row_idx in range(4, 200):
        for ci in [5,8]: ws.cell(row=row_idx, column=ci).fill = FILL(TEAL)
        for ci in [6,9]: ws.cell(row=row_idx, column=ci).fill = FILL(LT_BLUE)

    demand = pref_mb_demand(scouts)
    aquatics = {"Swimming","Lifesaving"}
    aq_staff_actual = 0; aq_staff_175 = 0
    dr = 4
    for mb, cnt in sorted(demand.items(), key=lambda x: -x[1]):
        disc  = is_discontinued(mb)
        eagle = is_eagle(mb)
        two_p = is_two_period(mb)
        aq    = mb in aquatics
        periods = 2 if two_p else 1
        if disc:
            sects_a=sects_175=staff_a=staff_175='N/A'
            notes = "⚠ DISCONTINUED BADGE — do not schedule"
            bg = LT_RED
        elif aq:
            sects_a   = ceil(cnt/12)
            sects_175 = ceil(TARGET*(cnt/total)/12) if total else 0
            staff_a   = '(aq. team)'; staff_175 = '(shared)'
            aq_staff_actual += sects_a; aq_staff_175 += sects_175
            notes = "Aquatics — staff shared across Swimming & Lifesaving"
            bg = LT_GREEN if eagle else (WHITE if (dr-4)%2==0 else ALT_GRAY)
        else:
            sects_a   = ceil(cnt/12)
            sects_175 = ceil(TARGET*(cnt/total)/12) if total else 0
            staff_a   = sects_a * 5; staff_175 = sects_175 * 5
            notes = "2-period badge (needs consecutive blocks)" if two_p else ""
            bg = LT_RED if disc else (LT_GREEN if eagle else (WHITE if (dr-4)%2==0 else ALT_GRAY))

        for ci, val in enumerate([mb, cnt, "Yes" if eagle else "No", periods,
                                   sects_a, sects_175, 5 if not (disc or aq) else ('N/A' if disc else '(shared)'),
                                   staff_a, staff_175, notes], 1):
            cell_bg = bg
            if isinstance(val, str) and ci == 5: cell_bg = TEAL
            if isinstance(val, str) and ci == 8: cell_bg = TEAL
            if isinstance(val, str) and ci == 6: cell_bg = LT_BLUE
            if isinstance(val, str) and ci == 9: cell_bg = LT_BLUE
            set_cell(ws, dr, ci, val, bg=cell_bg, halign='center' if ci>1 else 'left',
                     bold=disc, fg='8B0000' if disc else '000000')
        dr += 1

    # Aquatics note
    ws.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=10)
    ac = ws.cell(row=dr, column=1,
                 value=f"Aquatics team: {aq_staff_actual} staff shared across Swimming and Lifesaving (actual) / {aq_staff_175} @ 175.")
    ac.fill=FILL(TAN); ac.font=FONT(color='000000', size=9, italic=True)
    ws.row_dimensions[dr].height = 14; dr += 1

    # Total staff footer
    tot_actual  = sum(ceil(c/12)*5 for mb, c in demand.items() if not is_discontinued(mb) and mb not in aquatics)
    tot_175     = sum(ceil(TARGET*(c/total)/12)*5 for mb, c in demand.items()
                      if not is_discontinued(mb) and mb not in aquatics and total)
    for ci, val in enumerate(["TOTAL STAFF","","","","","","",tot_actual,tot_175,""], 1):
        sc = set_cell(ws, dr, ci, val, bg=MED_BLUE if ci==8 else (BLUE if ci==9 else GRAY),
                      fg=WHITE, bold=True, halign='center' if ci>1 else 'left')
    ws.row_dimensions[dr].height = 20

    col_widths(ws, [36,16,14,10,18,18,14,18,18,44])
    ws.freeze_panes = 'A4'

# ── TAB 4: DISCONTINUED BADGE OUTREACH ────────────────────────────────────────
def build_discontinued(wb, scouts):
    ws = wb.create_sheet("Discontinued Badge Outreach"); ws.sheet_properties.tabColor = RED
    ws.merge_cells('A1:J1')
    c = ws['A1']; c.value = "Discontinued Badge Outreach  |  Eagle Achievement Camp 2026"
    c.fill=FILL(RED); c.font=FONT(bold=True,size=13); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:J2')
    nc = ws['A2']
    nc.value = ("⚠ The badges below are discontinued by Scouting America. "
                "Please contact families to select an alternative.")
    nc.fill=FILL(LT_RED); nc.font=Font(name="Arial", italic=True, size=9, bold=True, color='8B0000')
    nc.alignment = ALIGN(h='left'); ws.row_dimensions[2].height = 14

    hdr(ws, 3, ["#","Scout Name","Unit","District","Discontinued Badge","Parent Name",
                "Parent Email","Parent Phone","Reg #","Contacted?"],
        RED, size=9, height=16)

    disc_scouts = [(s, [mb for mb in s.pref_mbs if is_discontinued(mb)])
                   for s in scouts if any(is_discontinued(mb) for mb in s.pref_mbs)]
    disc_scouts.sort(key=lambda x: x[0].name)
    for i, (s, disc_mbs) in enumerate(disc_scouts, 1):
        bg = WHITE if i%2==0 else LT_RED
        row = i+3
        for ci, val in enumerate([i, s.full_name, s.unit, s.district,
                                   "; ".join(disc_mbs), s.parent_name,
                                   s.parent_email, s.parent_phone, s.reg_num, ""], 1):
            set_cell(ws, row, ci, val, bg=bg)
    col_widths(ws, [5,26,30,18,36,22,28,16,14,12])
    ws.freeze_panes = 'A4'

# ── TAB 5: BALANCE DUE ────────────────────────────────────────────────────────
def build_balance_due(wb, scouts):
    ws = wb.create_sheet("Balance Due"); ws.sheet_properties.tabColor = ORANGE
    ws.merge_cells('A1:I1')
    c = ws['A1']; c.value = "Balance Due  |  Eagle Achievement Camp 2026"
    c.fill=FILL(ORANGE); c.font=FONT(bold=True,size=13); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 22
    hdr(ws, 2, ["#","Scout Name","Unit","District","Bus","Balance Due ($)","Reg #","Parent Email","Parent Phone"],
        ORANGE, size=9, height=16)

    bal_scouts = sorted([s for s in scouts if s.balance > 0], key=lambda s: -s.balance)
    for i, s in enumerate(bal_scouts, 1):
        bg = WHITE if i%2==0 else LT_YELLOW
        row = i+2
        for ci, val in enumerate([i, s.full_name, s.unit, s.district,
                                   'Yes' if s.bus=='1' else '', s.balance,
                                   s.reg_num, s.parent_email, s.parent_phone], 1):
            sc = set_cell(ws, row, ci, val, bg=bg, halign='center' if ci in (1,5,6) else 'left')
            if ci == 6: sc.number_format = '"$"#,##0.00'

    # Total row
    total_row = len(bal_scouts)+3
    ws.cell(row=total_row, column=1, value="TOTAL").fill = FILL(ORANGE)
    ws.cell(row=total_row, column=1).font = FONT(bold=True, color=WHITE, size=10)
    total_cell = ws.cell(row=total_row, column=6,
                         value=f"=SUM(F3:F{total_row-1})")
    total_cell.fill=FILL(ORANGE); total_cell.font=FONT(bold=True, color=WHITE, size=10)
    total_cell.number_format = '"$"#,##0.00'
    col_widths(ws, [5,26,30,18,6,14,13,28,16])
    ws.freeze_panes = 'A3'

# ── TAB 6: CAMPDOC STATUS ─────────────────────────────────────────────────────
def build_campdoc(wb, scouts):
    ws = wb.create_sheet("CampDoc Status"); ws.sheet_properties.tabColor = PURPLE
    ws.merge_cells('A1:G1')
    c = ws['A1']; c.value = "CampDoc Status  |  Eagle Achievement Camp 2026"
    c.fill=FILL(PURPLE); c.font=FONT(bold=True,size=13); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 22
    hdr(ws, 2, ["#","Scout Name","Unit","District","CampDoc Email","CampDoc %","Balance Due"],
        PURPLE, size=9, height=16)

    def cd_sort_key(s):
        if not s.cd_email: return (2, 0)
        if not s.cd_pct: return (1, 0)
        return (1 if s.cd_pct!='100' else 0, -(int(s.cd_pct) if s.cd_pct.isdigit() else 0))

    sorted_scouts = sorted(scouts, key=cd_sort_key)
    for i, s in enumerate(sorted_scouts, 1):
        row = i+2
        if not s.cd_email:                   bg = 'F0F0F0'
        elif s.cd_pct == '100':              bg = LT_GREEN
        elif s.cd_pct in ('', '0'):          bg = LT_RED
        else:                                bg = LT_YELLOW
        cd_val = (int(s.cd_pct) if s.cd_pct.isdigit() else '') if s.cd_email else ''
        for ci, val in enumerate([i, s.full_name, s.unit, s.district,
                                   s.cd_email, cd_val,
                                   s.balance if s.balance else ''], 1):
            sc = set_cell(ws, row, ci, val, bg=bg, halign='center' if ci in (1,6) else 'left')
            if ci == 6 and isinstance(val, int): sc.number_format = '0"%"'
            if ci == 7 and isinstance(val, float): sc.number_format = '"$"#,##0.00'
    col_widths(ws, [5,26,30,18,28,10,12])
    ws.freeze_panes = 'A3'

# ── TAB 7: SCOUT SCHEDULE ─────────────────────────────────────────────────────
def build_scout_schedule(wb, scouts, section_assignments):
    ws = wb.create_sheet("Scout Schedule"); ws.sheet_properties.tabColor = BLUE
    sched_lookup = scout_schedule(scouts, section_assignments)

    ws.merge_cells('A1:N1')
    c = ws['A1']
    c.value = "Camp Hi-Sierra  |  Scout Class Schedule  (Block Assignments from Registration)"
    c.fill=FILL(BLUE); c.font=FONT(bold=True,size=13); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 22

    # Legend
    ws.merge_cells('A2:N2')
    lc = ws['A2']
    lc.value = ("Green = all 5 blocks filled  |  Yellow = 4 filled  |  "
                "Orange = 3 filled  |  Red = 0-2 filled  |  Blue = 2-period badge")
    lc.fill=FILL('F9F9F9'); lc.font=Font(name="Arial", italic=True, size=9, color='444444')
    ws.row_dimensions[2].height = 14

    col_hdrs = ["#","Scout Name","Unit","District",
                "Block 1","Block 2","Block 3","Block 4","Block 5",
                "Placed","Prefs","Unplaced","Pref Summary","Notes"]
    hdr(ws, 3, col_hdrs, GRAY, size=10, height=36, wrap=True)
    ws.row_dimensions[3].height = 36

    # Sort by unit (SVMBC first) then name
    def sort_key(s):
        is_ooc = 0 if 'Silicon Valley' in s.council else 1
        return (is_ooc, s.unit, s.name)

    sorted_scouts = sorted(scouts, key=sort_key)

    for i, s in enumerate(sorted_scouts, 1):
        row = i+3
        ws.row_dimensions[row].height = 36
        sched = sched_lookup.get(id(s), [(None,None)]*5)
        filled = sum(1 for mb,_ in sched if mb is not None)
        prefs  = s.pref_mbs
        pref_count = len(prefs)
        placed_prefs = sum(1 for p in prefs if any(mb and p.lower() in mb.lower() for mb,_ in sched))
        unplaced = [p for p in prefs if not any(mb and p.lower() in mb.lower() for mb,_ in sched)]

        if   filled == 5:    row_bg = LT_GREEN
        elif filled == 4:    row_bg = LT_YELLOW
        elif filled == 3:    row_bg = 'FFD9B3'
        else:                row_bg = LT_RED

        pref_summary = f"Pref: {placed_prefs}/{pref_count}" if pref_count else "No prefs"

        for ci, val in enumerate([i, s.full_name, s.unit, s.district], 1):
            set_cell(ws, row, ci, val, bg=row_bg, halign='center' if ci==1 else 'left')

        # Block cells (columns 5-9)
        for bi, (mb, sec) in enumerate(sched):
            ci = bi + 5
            if mb is None:
                set_cell(ws, row, ci, "—", bg=LT_GRAY, halign='center')
            elif is_two_period(mb) and bi < 4 and sched[bi+1][0] and sched[bi+1][0] == mb:
                # First of 2-period pair
                set_cell(ws, row, ci, f"{mb}\nSec {sec} ★", bg=LT_BLUE,
                         fg=BLUE, bold=True, halign='center', wrap=True, size=9)
            elif is_two_period(mb) and bi > 0 and sched[bi-1][0] and sched[bi-1][0] == mb:
                # Continuation of 2-period
                set_cell(ws, row, ci, f"  ↳ {mb}\n  (cont. sec {sec})", bg=LT_BLUE2,
                         fg=BLUE, halign='center', wrap=True, size=9)
            else:
                set_cell(ws, row, ci, f"{mb}\nSec {sec}", bg=row_bg,
                         halign='center', wrap=True, size=9)

        for ci, val in enumerate([f"{filled}/5", pref_count,
                                   "; ".join(unplaced[:2])+"..." if len(unplaced)>2 else "; ".join(unplaced),
                                   pref_summary, s.admin_comment], 1):
            set_cell(ws, row, ci+9, val, bg=row_bg, halign='center' if ci<=2 else 'left',
                     bold=(ci==1), wrap=(ci>=3), size=9)

    # Summary footer
    placed_all  = sum(1 for s in sorted_scouts if sum(1 for mb,_ in sched_lookup.get(id(s),[]*(5)) if mb) == 5)
    placed_4    = sum(1 for s in sorted_scouts if sum(1 for mb,_ in sched_lookup.get(id(s),[]*(5)) if mb) == 4)
    placed_3    = sum(1 for s in sorted_scouts if sum(1 for mb,_ in sched_lookup.get(id(s),[]*(5)) if mb) == 3)
    placed_low  = sum(1 for s in sorted_scouts if sum(1 for mb,_ in sched_lookup.get(id(s),[]*(5)) if mb) < 3)
    foot_row = len(sorted_scouts)+4
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=14)
    fc = ws.cell(row=foot_row, column=1,
                 value=f"SUMMARY  |  {len(scouts)} scouts  |  "
                       f"All 5 filled: {placed_all}  |  4 filled: {placed_4}  |  "
                       f"3 filled: {placed_3}  |  ≤2 filled: {placed_low}")
    fc.fill=FILL(BLUE); fc.font=FONT(bold=True); fc.alignment=ALIGN(h='center')
    ws.row_dimensions[foot_row].height = 20

    col_widths(ws, [5,26,30,14,22,22,22,22,22,11,14,32,11,13])
    ws.freeze_panes = 'A4'

# ── TAB 8: CLASS ROSTERS ──────────────────────────────────────────────────────
def build_class_rosters(wb, scouts, section_assignments):
    ws = wb.create_sheet("Class Rosters"); ws.sheet_properties.tabColor = DK_GREEN
    ws.merge_cells('A1:H1')
    c = ws['A1']; c.value = "Eagle Achievement Camp 2026  |  Class Rosters by Section"
    c.fill=FILL(DK_GREEN); c.font=FONT(bold=True,size=13); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 22

    # Build ordered sections: {(badge, sec): [(scout, block_idx), ...]}
    sec_map = defaultdict(list)
    for (mb, bi), entries in section_assignments.items():
        for (s, sec_num) in entries:
            sec_map[(mb, sec_num)].append((s, bi))

    cur_row = 2
    disc_flag = {id(s) for s in scouts if any(is_discontinued(mb) for mb in s.pref_mbs)}

    for (mb, sec_num) in sorted(sec_map.keys(), key=lambda x: (x[0], x[1])):
        entries = sec_map[(mb, sec_num)]
        # Determine block (majority vote)
        block_votes = Counter(bi for _,bi in entries)
        block_idx = block_votes.most_common(1)[0][0]
        enrolled = len(entries)
        eagle = is_eagle(mb)

        # Section header
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=8)
        sh = ws.cell(row=cur_row, column=1,
                     value=f"  {mb}  |  Section {sec_num}  |  Block {block_idx+1}  |  {enrolled}/12 scouts")
        sh.fill=FILL(DK_GREEN if eagle else GRAY)
        sh.font=FONT(bold=True, size=11); sh.alignment=ALIGN(h='left')
        ws.row_dimensions[cur_row].height = 20; cur_row += 1

        # Column headers
        hdr(ws, cur_row, ["#","Scout Name","Unit","District","Rank","Bus","Pref #","Disc?"],
            GRAY, size=9, height=14)
        cur_row += 1

        scout_entries = sorted(entries, key=lambda x: (x[0].unit, x[0].name))
        for ri, (s, _) in enumerate(scout_entries, 1):
            bg = LT_GREEN if ri%2==0 else WHITE
            is_disc = id(s) in disc_flag
            pref_rank = next((str(j+1) for j,mb2 in enumerate(s.pref_mbs)
                              if mb2 and mb.lower() in mb2.lower()), '')
            for ci, val in enumerate([ri, s.full_name, s.unit, s.district,
                                       s.rank, 'Yes' if s.bus=='1' else '',
                                       pref_rank, "YES" if is_disc else ""], 1):
                sc = set_cell(ws, cur_row, ci, val, bg=LT_RED if (ci==8 and is_disc) else bg,
                              halign='center' if ci in (1,5,6,7,8) else 'left', size=9)
            cur_row += 1

        cur_row += 1  # blank separator

    col_widths(ws, [5,26,36,14,14,6,8,8])
    ws.freeze_panes = 'A2'

# ── TAB 9: BLOCK VIEW ─────────────────────────────────────────────────────────
def build_block_view(wb, scouts, section_assignments):
    ws = wb.create_sheet("Block View"); ws.sheet_properties.tabColor = MED_BLUE
    ws.merge_cells('A1:H1')
    c = ws['A1']; c.value = "Eagle Achievement Camp 2026  |  Block Grid View"
    c.fill=FILL(MED_BLUE); c.font=FONT(bold=True,size=13); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 22

    ws.cell(row=2, column=1, value="Badge / Section").fill = FILL(GRAY)
    ws.cell(row=2, column=1).font = FONT(bold=True, size=9)
    for bi in range(5):
        c = ws.cell(row=2, column=bi+2, value=f"Block {bi+1}")
        c.fill=FILL(BLUE); c.font=FONT(bold=True,size=10); c.alignment=ALIGN(h='center')
    c7 = ws.cell(row=2, column=7, value="Total\nEnrolled")
    c7.fill=FILL(DK_GREEN); c7.font=FONT(bold=True,size=9); c7.alignment=ALIGN(h='center',wrap=True)
    c8 = ws.cell(row=2, column=8, value="NOTES / Changes")
    c8.fill=FILL(BLUE); c8.font=FONT(bold=True,size=9)
    ws.row_dimensions[2].height = 28

    # Organize by badge/section/block
    section_block = defaultdict(dict)  # {(badge,sec): {block_idx: count}}
    for (mb, bi), entries in section_assignments.items():
        for (s, sec_num) in entries:
            if bi not in section_block[(mb, sec_num)]:
                section_block[(mb, sec_num)][bi] = 0
            section_block[(mb, sec_num)][bi] += 1

    cur_row = 3
    for ri, (mb, sec_num) in enumerate(sorted(section_block.keys(), key=lambda x:(x[0],x[1]))):
        block_counts = section_block[(mb, sec_num)]
        total_enrolled = sum(block_counts.values())
        two_p = is_two_period(mb)
        bg = LT_BLUE if (two_p and ri%2==0) else (WHITE if ri%2==0 else 'E2EFDA')

        label = f"{mb}  Sec {sec_num}" + (" ★" if two_p else "")
        a_cell = ws.cell(row=cur_row, column=1, value=label)
        a_cell.fill = FILL(LT_BLUE if two_p else bg)
        a_cell.font = Font(name="Arial", size=9, color='000000', bold=two_p)
        a_cell.alignment = ALIGN()
        a_cell.border = BORDER()

        for bi in range(5):
            cell = ws.cell(row=cur_row, column=bi+2)
            cell.border = BORDER()
            if bi in block_counts:
                cnt = block_counts[bi]
                cell.value = f"{cnt}/12"
                cell.alignment = ALIGN(h='center')
                if cnt >= 12:   cap_bg = 'FFC7CE'
                elif cnt >= 10: cap_bg = 'FFFF00'
                else:           cap_bg = 'C6EFCE'
                cell.fill = FILL(cap_bg)
                cell.font = Font(name="Arial", size=9)
            else:
                cell.fill = FILL(bg)

        tc = ws.cell(row=cur_row, column=7, value=total_enrolled)
        tc.fill=FILL(LT_GREEN); tc.font=Font(name="Arial",size=9); tc.alignment=ALIGN(h='center')
        tc.border=BORDER()
        ws.cell(row=cur_row, column=8).border=BORDER()
        cur_row += 1

    col_widths(ws, [40,12,12,12,12,12,12,30])
    ws.freeze_panes = 'A3'

# ── TAB 10: UNIT COHESION ─────────────────────────────────────────────────────
def build_unit_cohesion(wb, scouts, section_assignments):
    ws = wb.create_sheet("Unit Cohesion"); ws.sheet_properties.tabColor = ORANGE
    ws.merge_cells('A1:G1')
    c = ws['A1']; c.value = "Eagle Achievement Camp 2026  |  Unit Cohesion Analysis"
    c.fill=FILL(ORANGE); c.font=FONT(bold=True,size=13); c.alignment=ALIGN(h='center')
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:G2')
    nc = ws['A2']
    nc.value = "Shows how many classmates from the same unit are scheduled together per class. Higher = more cohesion."
    nc.font = Font(name="Arial", italic=True, size=9, color='000000')
    nc.alignment = ALIGN()
    ws.row_dimensions[2].height = 14

    hdr(ws, 3, ["Unit","# Scouts","Avg Classes Placed","Avg Unitmates Per Class",
                "Scouts w/ All 5 Blocks","Scouts w/ 4 Blocks","Scouts w/ ≤3 Blocks"],
        GRAY, size=9, height=28, wrap=True)

    # Build section membership: {(mb, bi, sec): set of scout ids}
    sec_members = defaultdict(set)
    for (mb, bi), entries in section_assignments.items():
        for (s, sec_num) in entries:
            sec_members[(mb, bi, sec_num)].add(id(s))

    by_unit = defaultdict(list)
    for s in scouts: by_unit[s.unit or '(No Unit)'].append(s)

    unit_rows = []
    for unit, unit_scouts in by_unit.items():
        unit_ids = {id(s) for s in unit_scouts}
        n = len(unit_scouts)
        all_5 = sum(1 for s in unit_scouts
                    if sum(1 for raw in s.blocks if any(normalize_mb(raw))) == 5)
        four  = sum(1 for s in unit_scouts
                    if sum(1 for raw in s.blocks if any(normalize_mb(raw))) == 4)
        low   = n - all_5 - four

        # Avg classes placed
        placed_counts = []
        for s in unit_scouts:
            placed_counts.append(sum(1 for raw in s.blocks if any(normalize_mb(raw))))
        avg_placed = round(sum(placed_counts)/n, 1) if n else 0

        # Avg unitmates per class
        unitmate_sums = []
        for s in unit_scouts:
            for bi, raw in enumerate(s.blocks):
                for mb in normalize_mb(raw):
                    # find section
                    for (mb2, bi2, sec_num), members in sec_members.items():
                        if mb2 == mb and bi2 == bi and id(s) in members:
                            unitmates = len(members & unit_ids) - 1
                            unitmate_sums.append(unitmates)
                            break
        avg_unitmates = round(sum(unitmate_sums)/len(unitmate_sums), 1) if unitmate_sums else 0

        unit_rows.append((unit, n, avg_placed, avg_unitmates, all_5, four, low))

    unit_rows.sort(key=lambda x: (-x[1], x[0]))
    for ri, row in enumerate(unit_rows, 1):
        bg = WHITE if ri%2==0 else ALT_GRAY
        for ci, val in enumerate(row, 1):
            set_cell(ws, ri+3, ci, val, bg=bg, halign='center' if ci>1 else 'left', size=9)

    col_widths(ws, [40,10,16,18,18,14,14])
    ws.freeze_panes = 'A4'

# ── FIRESTORE PUSH ────────────────────────────────────────────────────────────
def push_to_firestore(scouts, service_account_path):
    """Push mbDemand + session stats to Firestore sessions/Eagle_Achievement_Camp."""
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore as fs
    except ImportError:
        print("⚠  firebase-admin not installed. Run: pip install firebase-admin --break-system-packages")
        return

    try:
        app = firebase_admin.get_app()
    except ValueError:
        cred = credentials.Certificate(service_account_path)
        app = firebase_admin.initialize_app(cred)

    db = fs.client()
    doc_id = 'Eagle_Achievement_Camp'

    total   = len(scouts)
    svmbc   = sum(1 for s in scouts if 'Silicon Valley' in s.council)
    units   = sorted({s.unit for s in scouts if s.unit})

    # Use block assignments (actual schedule) for demand; fall back to prefs
    mb_demand = block_mb_demand(scouts)
    if not mb_demand:
        mb_demand = pref_mb_demand(scouts)

    cd_nums     = [int(s.cd_pct) for s in scouts if s.cd_pct.isdigit() and s.cd_email]
    avg_cd      = round(sum(cd_nums) / len(cd_nums), 1) if cd_nums else 0
    cd_complete = sum(1 for s in scouts if s.cd_pct == '100')
    cd_started  = sum(1 for s in scouts if s.cd_email and s.cd_pct not in ('', '0'))
    cd_total    = sum(1 for s in scouts if s.cd_email)
    balance_due = round(sum(s.balance for s in scouts if s.balance > 0), 2)

    doc_data = {
        'sessionName':          'Eagle Achievement Camp',
        'type':                 'eagle_camp',
        'youthCount':           total,
        'totalRegistrants':     total,
        'totalCount':           total,
        'adultCount':           0,
        'unitCount':            len(units),
        'svmbcCount':           svmbc,
        'outOfCouncilCount':    total - svmbc,
        'mbDemand':             dict(mb_demand),
        'balanceDue':           balance_due,
        'totalBalanceDue':      balance_due,
        'campdocTotalCount':    cd_total,
        'campdocStartedCount':  cd_started,
        'campdocCompleteCount': cd_complete,
        'avgCampdocPct':        avg_cd,
        'lastSync':             int(datetime.utcnow().timestamp() * 1000),
        'lastSyncSource':       'BlackPug_EventDataDump',
    }

    print(f"  [{doc_id}] Writing session metadata...", end=' ', flush=True)
    try:
        db.collection('sessions').document(doc_id).set(doc_data, merge=True)
        print("✅")
    except Exception as e:
        print(f"❌ Error: {e}")
        return

    if mb_demand:
        print(f"  [{doc_id}] Writing mbDemand subcollection...", end=' ', flush=True)
        try:
            db.collection('sessions').document(doc_id)\
              .collection('mbDemand').document('current')\
              .set({'demand': dict(mb_demand), 'updatedAt': fs.SERVER_TIMESTAMP}, merge=True)
            print("✅")
        except Exception as e:
            print(f"❌ Error: {e}")

    print("Firestore push complete — Eagle_Achievement_Camp updated.")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]
    service_account = None
    push_only = False
    out_dir = "."

    if '--service-account' in args:
        idx = args.index('--service-account')
        service_account = args[idx + 1]; args = args[:idx] + args[idx + 2:]
    if '--push-only' in args:
        idx = args.index('--push-only')
        service_account = args[idx + 1]; args = args[:idx] + args[idx + 2:]
        push_only = True
    if len(args) > 1 and not args[-1].lower().endswith('.csv'):
        out_dir = args[-1]; args = args[:-1]

    csv_path = args[0] if args else None
    if not csv_path:
        print("Usage: python build_eac_workbook.py <Event_Data_Dump.csv> [output_dir]")
        print("       Add --service-account <key.json> to also push to Firestore")
        print("       Add --push-only <key.json> to skip Excel and only push to Firestore")
        sys.exit(1)

    print(f"Loading {os.path.basename(csv_path)}...")
    scouts, event_name, session_start, unnamed_regs = load_csv(csv_path)

    print(f"  Scouts with names: {len(scouts)}")
    if unnamed_regs:
        print(f"  ⚠ {len(unnamed_regs)} registration row(s) with empty First Name:")
        for r in unnamed_regs:
            print(f"    Reg# {r.get('Registration Number','')}  Type: {r.get('Registrant Type','')}  PID: {r.get('Participant ID','')}")

    # ── Firestore push ──
    if service_account:
        print(f"\nPushing to Firestore...")
        push_to_firestore(scouts, service_account)

    if push_only:
        return

    print("Building section assignments from block data...")
    sa = build_section_assignments(scouts)
    badges_with_sections = len(set(mb for mb, _ in sa.keys()))
    print(f"  {badges_with_sections} badge classes across {len(sa)} badge/block slots")

    wb = Workbook(); wb.remove(wb.active)

    print("Tab 1: Dashboard..."); build_dashboard(wb, scouts, event_name, session_start)
    print("Tab 2: Scout Roster..."); build_scout_roster(wb, scouts)
    print("Tab 3: MB Scheduling..."); build_mb_scheduling(wb, scouts, event_name)
    print("Tab 4: Discontinued Badge Outreach..."); build_discontinued(wb, scouts)
    print("Tab 5: Balance Due..."); build_balance_due(wb, scouts)
    print("Tab 6: CampDoc Status..."); build_campdoc(wb, scouts)
    print("Tab 7: Scout Schedule..."); build_scout_schedule(wb, scouts, sa)
    print("Tab 8: Class Rosters..."); build_class_rosters(wb, scouts, sa)
    print("Tab 9: Block View..."); build_block_view(wb, scouts, sa)
    print("Tab 10: Unit Cohesion..."); build_unit_cohesion(wb, scouts, sa)

    today = date.today().strftime("%Y%m%d")
    out_path = os.path.join(out_dir, f"EAC_2026_Scheduling_{today}.xlsx")
    wb.save(out_path)
    print(f"\n✅ Saved: {out_path}")

    # Post-output summary
    disc_scouts = [s for s in scouts if any(is_discontinued(mb) for mb in s.pref_mbs)]
    bal_scouts  = [s for s in scouts if s.balance > 0]
    cd_100  = sum(1 for s in scouts if s.cd_pct=='100')
    cd_part = sum(1 for s in scouts if s.cd_email and s.cd_pct not in ('','0','100'))
    cd_zero = sum(1 for s in scouts if s.cd_pct in ('0','') and s.cd_email)
    cd_none = sum(1 for s in scouts if not s.cd_email)
    demand  = pref_mb_demand(scouts)
    print(f"\n── POST-OUTPUT SUMMARY ─────────────────────────────")
    print(f"Total scouts processed: {len(scouts)}")
    print(f"Scouts w/ discontinued badge preference: {len(disc_scouts)}")
    if disc_scouts:
        disc_mbs = Counter(mb for s in disc_scouts for mb in s.pref_mbs if is_discontinued(mb))
        for mb, c in disc_mbs.items(): print(f"  - {mb}: {c} scouts")
    print(f"Balance due: {len(bal_scouts)} scouts  /  ${sum(s.balance for s in bal_scouts):,.2f} total")
    print(f"CampDoc: {cd_100} complete  /  {cd_part} partial  /  {cd_zero} not started  /  {cd_none} no email")
    # Duplicate names
    names = Counter(s.full_name for s in scouts)
    dups = [(n,c) for n,c in names.items() if c>1]
    if dups: print(f"Duplicate names: {dups}")
    else: print("No duplicate names.")
    # Schedule placement
    from collections import Counter as C2
    fill_counts = C2(sum(1 for raw in s.blocks if any(normalize_mb(raw))) for s in scouts)
    print(f"Block fill: all 5: {fill_counts[5]}  |  4: {fill_counts[4]}  |  3: {fill_counts[3]}  |  ≤2: {sum(v for k,v in fill_counts.items() if k<3)}")

if __name__ == '__main__':
    main()
